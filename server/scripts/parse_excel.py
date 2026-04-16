#!/usr/bin/env python3
"""
Robust Excel parser for NuPhorm biostatistics platform.
Uses pandas + openpyxl for reliable parsing of complex
research spreadsheets with merged cells, multi-row headers,
and non-standard formatting.

Usage: python parse_excel.py <filepath> [--sheet <name>]
Output: JSON to stdout
"""

import sys
import json
import pandas as pd
import openpyxl
import warnings
import os

warnings.filterwarnings('ignore', category=UserWarning)


def detect_header_row(filepath, sheet_name=None, max_scan=15):
    """
    Scan the first N rows to find the actual header row.
    The header row is the first row where >40% of cells contain
    non-empty string-like values (not just numbers).
    """
    wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    best_row = 0
    best_score = 0
    max_col = ws.max_column or 1

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=max_scan, values_only=True)):
        if row is None:
            continue

        filled = 0
        string_like = 0
        total = min(len(row), max_col)
        filled_values = []

        for cell_val in row[:total]:
            if cell_val is not None and str(cell_val).strip() != '':
                filled += 1
                filled_values.append(str(cell_val).strip())
                if isinstance(cell_val, str) and not cell_val.replace('.', '').replace('-', '').isdigit():
                    string_like += 1

        if total == 0 or filled == 0:
            continue

        # Uniqueness: a former merged-title row will have all cells equal after unmerge.
        # Real header rows have mostly-unique column names.
        uniqueness_ratio = len(set(filled_values)) / filled
        if filled >= 2 and uniqueness_ratio < 0.6:
            continue

        fill_ratio = filled / total
        string_ratio = string_like / total
        score = (fill_ratio * 0.3) + (string_ratio * 0.4) + (uniqueness_ratio * 0.3)

        if score > best_score and fill_ratio > 0.3:
            best_score = score
            best_row = row_idx

    wb.close()
    return best_row


def unmerge_and_parse(filepath, sheet_name=None):
    """
    Load workbook, handle merged cells by filling them with
    the merged region's value, detect the header row, then
    parse with pandas.
    """
    wb = openpyxl.load_workbook(filepath, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active

    merged_ranges = list(ws.merged_cells.ranges)
    merge_count = len(merged_ranges)

    for merged_range in merged_ranges:
        min_row, min_col = merged_range.min_row, merged_range.min_col
        top_left_value = ws.cell(row=min_row, column=min_col).value
        ws.unmerge_cells(str(merged_range))
        for row in range(min_row, merged_range.max_row + 1):
            for col in range(min_col, merged_range.max_col + 1):
                ws.cell(row=row, column=col).value = top_left_value

    temp_path = filepath + '.tmp.xlsx'
    wb.save(temp_path)
    wb.close()

    return temp_path, merge_count


def parse_excel(filepath, sheet_name=None):
    """Main entry point: unmerge -> detect header -> parse -> clean -> output."""

    result = {
        'success': True,
        'rows': [],
        'columns': [],
        'csvText': '',
        'metadata': {
            'original_file': os.path.basename(filepath),
            'sheet_name': sheet_name,
            'merges_fixed': 0,
            'header_row': 1,
            'total_rows': 0,
            'total_columns': 0,
        },
    }

    try:
        # Step 1: Unmerge cells
        temp_path, merge_count = unmerge_and_parse(filepath, sheet_name)
        result['metadata']['merges_fixed'] = merge_count

        # Step 2: Detect header row (on the unmerged version)
        header_row = detect_header_row(temp_path, sheet_name)
        result['metadata']['header_row'] = header_row + 1

        # Step 3: Parse with pandas
        df = pd.read_excel(
            temp_path,
            sheet_name=sheet_name or 0,
            header=header_row,
            engine='openpyxl',
        )

        # Step 4: Clean up
        unnamed_cols = [c for c in df.columns if str(c).startswith('Unnamed:')]
        cols_to_drop = [c for c in unnamed_cols if df[c].isna().all()]
        df = df.drop(columns=cols_to_drop)
        df = df.dropna(how='all')
        df.columns = [str(c).strip().replace('\n', ' ').replace('\r', '') for c in df.columns]
        df = df.fillna('')

        # Step 5: Build output
        columns = df.columns.tolist()
        rows = df.to_dict(orient='records')

        # Build CSV text
        csv_text = df.to_csv(index=False)

        result['rows'] = rows
        result['columns'] = columns
        result['csvText'] = csv_text
        result['metadata']['total_rows'] = len(rows)
        result['metadata']['total_columns'] = len(columns)

        # Clean up temp file
        if os.path.exists(temp_path):
            os.remove(temp_path)

    except Exception as e:
        result['success'] = False
        result['error'] = str(e)
        temp_path = filepath + '.tmp.xlsx'
        if os.path.exists(temp_path):
            os.remove(temp_path)

    return result


if __name__ == '__main__':
    if len(sys.argv) < 2:
        print(json.dumps({'success': False, 'error': 'No file path provided'}))
        sys.exit(1)

    filepath = sys.argv[1]
    sheet_name = None

    if '--sheet' in sys.argv:
        idx = sys.argv.index('--sheet')
        if idx + 1 < len(sys.argv):
            sheet_name = sys.argv[idx + 1]

    result = parse_excel(filepath, sheet_name)
    print(json.dumps(result, default=str, ensure_ascii=False))
