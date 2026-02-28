"""
Extended analytics endpoints for NuPhorm Biostatistics Platform.

  POST /api/v1/stats/inferential   — t-tests, ANOVA, regression, non-parametric
  POST /api/v1/stats/sample-size   — power analysis / sample size calculation
  POST /api/v1/stats/script        — execute custom Python script
  POST /api/v1/data/upload         — parse CSV/XLSX/TSV file
  POST /api/v1/data/clean          — handle missing values, outliers, normalization
  POST /api/v1/data/transform      — log/sqrt/zscore/minmax/groupby transformations
"""
from __future__ import annotations

import csv
import io
import json
import math
import logging
import traceback
from contextlib import redirect_stdout
from io import StringIO
from typing import Any

import numpy as np
from fastapi import APIRouter, HTTPException, UploadFile, File, status
from pydantic import BaseModel
from scipy import stats as scipy_stats

logger = logging.getLogger(__name__)
router = APIRouter()


# ── Input models ──────────────────────────────────────────────────────────────

class InferentialRequest(BaseModel):
    test_type: str
    group1: list[float]
    group2: list[float] = []
    alpha: float = 0.05
    alternative: str = "two-sided"  # "two-sided", "less", "greater"
    mu0: float = 0.0  # hypothesised mean for one-sample tests


class SampleSizeRequest(BaseModel):
    test_type: str  # t-test-one-sample, t-test-two-sample, t-test-paired, ...
    effect_size: float
    alpha: float = 0.05
    power: float = 0.80
    ratio: float = 1.0  # allocation ratio n2/n1


class ScriptRequest(BaseModel):
    code: str
    data: list[dict[str, Any]] = []


class DataCleanRequest(BaseModel):
    rows: list[dict[str, Any]]
    missing_strategy: str = "mean"  # drop | mean | median | zero
    remove_outliers: bool = False
    normalize: bool = False


class DataTransformRequest(BaseModel):
    rows: list[dict[str, Any]]
    transform_type: str  # log | sqrt | zscore | minmax | group_by
    column: str


# ── Inferential statistics ────────────────────────────────────────────────────

@router.post(
    "/stats/inferential",
    summary="Run an inferential statistical test",
    status_code=status.HTTP_200_OK,
)
async def run_inferential(req: InferentialRequest) -> dict:
    """
    test_type values:
      independent_t, paired_t, one_sample_t,
      one_way_anova, two_way_anova,
      wilcoxon, mann_whitney, kruskal_wallis,
      linear_regression, logistic_regression, mixed_effects
    """
    g1 = np.array(req.group1)
    g2 = np.array(req.group2) if req.group2 else np.array([])
    alt = req.alternative

    try:
        if req.test_type == "independent_t":
            if len(g2) == 0:
                raise ValueError("group2 required for independent t-test")
            stat, p = scipy_stats.ttest_ind(g1, g2, alternative=alt)
            n1, n2 = len(g1), len(g2)
            pooled_std = np.sqrt(
                ((n1 - 1) * np.std(g1, ddof=1) ** 2 + (n2 - 1) * np.std(g2, ddof=1) ** 2)
                / (n1 + n2 - 2)
            )
            cohens_d = (np.mean(g1) - np.mean(g2)) / pooled_std if pooled_std > 0 else 0.0
            ci = scipy_stats.t.interval(
                1 - req.alpha,
                df=n1 + n2 - 2,
                loc=np.mean(g1) - np.mean(g2),
                scale=np.sqrt(np.var(g1, ddof=1) / n1 + np.var(g2, ddof=1) / n2),
            )
            return {
                "test": "Independent Samples t-Test",
                "statistic": round(float(stat), 4),
                "p_value": round(float(p), 4),
                "significant": bool(p < req.alpha),
                "cohens_d": round(float(cohens_d), 4),
                "ci_lower": round(float(ci[0]), 4),
                "ci_upper": round(float(ci[1]), 4),
                "mean_group1": round(float(np.mean(g1)), 4),
                "mean_group2": round(float(np.mean(g2)), 4),
                "n_group1": n1,
                "n_group2": n2,
            }

        elif req.test_type == "paired_t":
            if len(g2) == 0:
                raise ValueError("group2 required for paired t-test")
            stat, p = scipy_stats.ttest_rel(g1, g2, alternative=alt)
            diff = g1 - g2
            cohens_d = float(np.mean(diff) / np.std(diff, ddof=1)) if np.std(diff, ddof=1) > 0 else 0.0
            return {
                "test": "Paired Samples t-Test",
                "statistic": round(float(stat), 4),
                "p_value": round(float(p), 4),
                "significant": bool(p < req.alpha),
                "cohens_d": round(cohens_d, 4),
                "mean_difference": round(float(np.mean(diff)), 4),
                "n_pairs": len(g1),
            }

        elif req.test_type == "one_sample_t":
            stat, p = scipy_stats.ttest_1samp(g1, popmean=req.mu0, alternative=alt)
            cohens_d = float((np.mean(g1) - req.mu0) / np.std(g1, ddof=1)) if np.std(g1, ddof=1) > 0 else 0.0
            return {
                "test": "One-Sample t-Test",
                "statistic": round(float(stat), 4),
                "p_value": round(float(p), 4),
                "significant": bool(p < req.alpha),
                "cohens_d": round(cohens_d, 4),
                "sample_mean": round(float(np.mean(g1)), 4),
                "hypothesized_mean": req.mu0,
                "n": len(g1),
            }

        elif req.test_type == "one_way_anova":
            groups = [g1] + ([g2] if len(g2) > 0 else [])
            stat, p = scipy_stats.f_oneway(*groups)
            return {
                "test": "One-Way ANOVA",
                "f_statistic": round(float(stat), 4),
                "p_value": round(float(p), 4),
                "significant": bool(p < req.alpha),
                "n_groups": len(groups),
                "group_means": [round(float(np.mean(g)), 4) for g in groups],
            }

        elif req.test_type == "kruskal_wallis":
            groups = [g1] + ([g2] if len(g2) > 0 else [])
            stat, p = scipy_stats.kruskal(*groups)
            return {
                "test": "Kruskal-Wallis H-Test",
                "h_statistic": round(float(stat), 4),
                "p_value": round(float(p), 4),
                "significant": bool(p < req.alpha),
                "n_groups": len(groups),
            }

        elif req.test_type == "wilcoxon":
            if len(g2) > 0:
                stat, p = scipy_stats.wilcoxon(g1, g2, alternative=alt)
            else:
                stat, p = scipy_stats.wilcoxon(g1, alternative=alt)
            return {
                "test": "Wilcoxon Signed-Rank Test",
                "statistic": round(float(stat), 4),
                "p_value": round(float(p), 4),
                "significant": bool(p < req.alpha),
                "n": len(g1),
            }

        elif req.test_type == "mann_whitney":
            if len(g2) == 0:
                raise ValueError("group2 required for Mann-Whitney U test")
            stat, p = scipy_stats.mannwhitneyu(g1, g2, alternative=alt)
            return {
                "test": "Mann-Whitney U Test",
                "u_statistic": round(float(stat), 4),
                "p_value": round(float(p), 4),
                "significant": bool(p < req.alpha),
                "n_group1": len(g1),
                "n_group2": len(g2),
            }

        elif req.test_type == "linear_regression":
            if len(g2) == 0:
                raise ValueError("group2 (Y values) required for linear regression")
            slope, intercept, r_value, p_value, std_err = scipy_stats.linregress(g1, g2)
            return {
                "test": "Linear Regression",
                "slope": round(float(slope), 4),
                "intercept": round(float(intercept), 4),
                "r_squared": round(float(r_value ** 2), 4),
                "r": round(float(r_value), 4),
                "p_value": round(float(p_value), 4),
                "significant": bool(p_value < req.alpha),
                "std_error": round(float(std_err), 4),
                "n": len(g1),
            }

        elif req.test_type == "logistic_regression":
            try:
                import statsmodels.api as sm  # type: ignore[import]
                X = sm.add_constant(g1)
                model = sm.Logit(g2, X)
                result = model.fit(disp=0)
                return {
                    "test": "Logistic Regression",
                    "intercept": round(float(result.params[0]), 4),
                    "coefficient": round(float(result.params[1]), 4),
                    "odds_ratio": round(float(np.exp(result.params[1])), 4),
                    "p_value_coeff": round(float(result.pvalues[1]), 4),
                    "significant": bool(result.pvalues[1] < req.alpha),
                    "aic": round(float(result.aic), 2),
                    "pseudo_r2": round(float(result.prsquared), 4),
                    "n": int(result.nobs),
                }
            except ImportError:
                raise HTTPException(
                    status_code=501,
                    detail="statsmodels not installed. Run: pip install statsmodels",
                )

        elif req.test_type == "mixed_effects":
            try:
                import statsmodels.formula.api as smf  # type: ignore[import]
                import pandas as pd  # type: ignore[import]
                n = len(g1)
                group_ids = np.repeat([0, 1], [n // 2, n - n // 2])
                df = pd.DataFrame({"y": g1, "group": group_ids})
                model = smf.mixedlm("y ~ group", df, groups=df["group"])
                result = model.fit(disp=False)
                return {
                    "test": "Linear Mixed-Effects Model",
                    "intercept": round(float(result.params["Intercept"]), 4),
                    "group_effect": round(float(result.params.get("group", 0.0)), 4),
                    "log_likelihood": round(float(result.llf), 3),
                    "aic": round(float(result.aic), 2),
                    "n": n,
                }
            except ImportError:
                raise HTTPException(
                    status_code=501,
                    detail="statsmodels not installed. Run: pip install statsmodels",
                )

        else:
            raise HTTPException(status_code=400, detail=f"Unknown test_type: '{req.test_type}'")

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=422, detail=str(exc))


# ── Sample size / power analysis ──────────────────────────────────────────────

@router.post(
    "/stats/sample-size",
    summary="Power analysis — calculate required sample size",
    status_code=status.HTTP_200_OK,
)
async def calculate_sample_size(req: SampleSizeRequest) -> dict:
    """
    test_type values:
      t-test-one-sample, t-test-two-sample, t-test-paired,
      proportion-one-sample, proportion-two-sample, anova-one-way
    """
    try:
        from statsmodels.stats import power as sm_power  # type: ignore[import]

        if req.test_type in ("t-test-one-sample", "t-test-paired"):
            analysis = sm_power.TTestPower()
            n = analysis.solve_power(
                effect_size=req.effect_size,
                alpha=req.alpha,
                power=req.power,
                alternative="two-sided",
            )
            n_ceil = math.ceil(n)
            achieved = analysis.solve_power(
                effect_size=req.effect_size,
                nobs=n_ceil,
                alpha=req.alpha,
            )
            return {
                "test": req.test_type.replace("-", " ").title(),
                "n": n_ceil,
                "total_n": n_ceil,
                "effect_size_cohen_d": req.effect_size,
                "alpha": req.alpha,
                "target_power": req.power,
                "achieved_power": round(float(achieved), 3),
            }

        elif req.test_type == "t-test-two-sample":
            analysis = sm_power.TTestIndPower()
            n = analysis.solve_power(
                effect_size=req.effect_size,
                alpha=req.alpha,
                power=req.power,
                ratio=req.ratio,
                alternative="two-sided",
            )
            n1 = math.ceil(n)
            n2 = math.ceil(n * req.ratio)
            return {
                "test": "Two-Sample t-Test",
                "n_group1": n1,
                "n_group2": n2,
                "total_n": n1 + n2,
                "allocation_ratio": req.ratio,
                "effect_size_cohen_d": req.effect_size,
                "alpha": req.alpha,
                "target_power": req.power,
            }

        elif req.test_type in ("proportion-one-sample", "proportion-two-sample"):
            analysis = sm_power.NormalIndPower()
            n = analysis.solve_power(
                effect_size=req.effect_size,
                alpha=req.alpha,
                power=req.power,
            )
            if req.test_type == "proportion-two-sample":
                n1 = math.ceil(n)
                n2 = math.ceil(n * req.ratio)
                return {
                    "test": "Two-Sample Proportion Test",
                    "n_group1": n1,
                    "n_group2": n2,
                    "total_n": n1 + n2,
                    "effect_size_cohen_h": req.effect_size,
                    "alpha": req.alpha,
                    "target_power": req.power,
                }
            else:
                return {
                    "test": "One-Sample Proportion Test",
                    "n": math.ceil(n),
                    "total_n": math.ceil(n),
                    "effect_size_cohen_h": req.effect_size,
                    "alpha": req.alpha,
                    "target_power": req.power,
                }

        elif req.test_type == "anova-one-way":
            analysis = sm_power.FTestAnovaPower()
            n = analysis.solve_power(
                effect_size=req.effect_size,
                alpha=req.alpha,
                power=req.power,
                k_groups=3,
            )
            n_ceil = math.ceil(n)
            return {
                "test": "One-Way ANOVA (k=3 groups)",
                "n_per_group": n_ceil,
                "total_n": n_ceil * 3,
                "k_groups": 3,
                "effect_size_cohen_f": req.effect_size,
                "alpha": req.alpha,
                "target_power": req.power,
            }

        else:
            raise HTTPException(status_code=400, detail=f"Unknown test_type: '{req.test_type}'")

    except ImportError:
        # Fallback via normal approximation (no statsmodels)
        z_alpha = float(scipy_stats.norm.ppf(1.0 - req.alpha / 2.0))
        z_beta = float(scipy_stats.norm.ppf(req.power))
        n_approx = math.ceil(((z_alpha + z_beta) / max(req.effect_size, 1e-9)) ** 2)
        multiplier = 2 if "two" in req.test_type else 1
        return {
            "test": req.test_type.replace("-", " ").title(),
            "n_per_group": n_approx,
            "total_n": n_approx * multiplier,
            "effect_size": req.effect_size,
            "alpha": req.alpha,
            "target_power": req.power,
            "note": "statsmodels unavailable; result uses normal approximation — install statsmodels for accuracy",
        }
    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=422, detail=str(exc))


# ── Custom script execution ───────────────────────────────────────────────────

@router.post(
    "/stats/script",
    summary="Execute custom Python analysis script (sandboxed)",
    status_code=status.HTTP_200_OK,
)
async def run_script(req: ScriptRequest) -> dict:
    """
    Runs arbitrary Python code in a restricted namespace.
    Available names: np, math, json, scipy_stats, pd (if installed).
    The variable `data` contains the submitted rows as a list of dicts.
    Define a variable named `result` (dict/list/scalar) to surface it
    separately in the response alongside captured stdout.
    """
    safe_builtins: dict[str, Any] = {
        "print": print,
        "range": range, "len": len, "list": list, "dict": dict,
        "tuple": tuple, "set": set, "str": str, "int": int,
        "float": float, "bool": bool, "type": type,
        "enumerate": enumerate, "zip": zip, "map": map, "filter": filter,
        "sorted": sorted, "min": min, "max": max, "sum": sum,
        "abs": abs, "round": round, "isinstance": isinstance,
        "vars": vars, "hasattr": hasattr, "getattr": getattr,
    }
    safe_globals: dict[str, Any] = {
        "__builtins__": safe_builtins,
        "np": np,
        "math": math,
        "json": json,
        "scipy_stats": scipy_stats,
        "data": req.data,
    }
    try:
        import pandas as pd  # type: ignore[import]
        safe_globals["pd"] = pd
    except ImportError:
        pass

    stdout_buf = StringIO()
    try:
        with redirect_stdout(stdout_buf):
            exec(req.code, safe_globals)  # noqa: S102
        output = stdout_buf.getvalue()
        result_var = safe_globals.get("result")
        serialisable = isinstance(result_var, (dict, list, str, int, float, bool))
        return {
            "output": output,
            "result": result_var if serialisable else None,
            "status": "success",
        }
    except Exception as exc:
        tb_lines = traceback.format_exc().strip().split("\n")
        short_tb = "\n".join(tb_lines[-4:])
        raise HTTPException(
            status_code=422,
            detail=f"{type(exc).__name__}: {exc}\n\n{short_tb}",
        )


# ── Data upload ───────────────────────────────────────────────────────────────

@router.post(
    "/data/upload",
    summary="Parse an uploaded CSV / XLSX / TSV / SAS dataset",
    status_code=status.HTTP_200_OK,
)
async def upload_dataset(file: UploadFile = File(...)) -> dict:
    content = await file.read()
    filename = (file.filename or "").lower()

    try:
        rows: list[dict] = []

        if filename.endswith((".csv", ".tsv")):
            sep = "\t" if filename.endswith(".tsv") else ","
            reader = csv.DictReader(
                io.StringIO(content.decode("utf-8-sig")), delimiter=sep
            )
            for row in reader:
                rows.append({k: _coerce(v) for k, v in row.items()})

        elif filename.endswith((".xlsx", ".xls")):
            try:
                import openpyxl  # type: ignore[import]
                wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
                ws = wb.active
                headers = [str(c.value) for c in next(ws.iter_rows(max_row=1))]
                for r in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, r)))
            except ImportError:
                raise HTTPException(
                    status_code=501,
                    detail="openpyxl not installed. Run: pip install openpyxl",
                )

        elif filename.endswith((".sas7bdat", ".xpt")):
            try:
                import pandas as pd  # type: ignore[import]
                fmt = "sas7bdat" if filename.endswith(".sas7bdat") else "xport"
                df = pd.read_sas(io.BytesIO(content), format=fmt)
                rows = df.to_dict("records")
            except ImportError:
                raise HTTPException(
                    status_code=501,
                    detail="pandas not installed. Run: pip install pandas pyreadstat",
                )

        else:
            raise HTTPException(
                status_code=415,
                detail=f"Unsupported file type '{file.filename}'. Use CSV, XLSX, XLS, TSV, SAS7BDAT, or XPT.",
            )

        columns = list(rows[0].keys()) if rows else []
        return {
            "name": file.filename,
            "rows": rows,
            "columns": columns,
            "row_count": len(rows),
            "column_count": len(columns),
        }

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=422, detail=f"Failed to parse file: {exc}")


# ── Data clean ────────────────────────────────────────────────────────────────

@router.post(
    "/data/clean",
    summary="Clean dataset: missing values, outlier removal, normalisation",
    status_code=status.HTTP_200_OK,
)
async def clean_dataset(req: DataCleanRequest) -> dict:
    if not req.rows:
        raise HTTPException(status_code=422, detail="rows must not be empty")

    try:
        import pandas as pd  # type: ignore[import]

        df = pd.DataFrame(req.rows)
        original_count = len(df)
        audit: list[str] = []
        numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()

        # ── Missing values ─────────────────────────────────────────────────
        if req.missing_strategy == "drop":
            df = df.dropna()
            removed = original_count - len(df)
            audit.append(f"Dropped {removed} rows with any missing value")
        elif req.missing_strategy == "mean":
            for col in numeric_cols:
                n_missing = int(df[col].isna().sum())
                if n_missing:
                    m = df[col].mean()
                    df[col] = df[col].fillna(m)
                    audit.append(f"Imputed {n_missing} missing in '{col}' → mean ({m:.3f})")
        elif req.missing_strategy == "median":
            for col in numeric_cols:
                n_missing = int(df[col].isna().sum())
                if n_missing:
                    med = df[col].median()
                    df[col] = df[col].fillna(med)
                    audit.append(f"Imputed {n_missing} missing in '{col}' → median ({med:.3f})")
        elif req.missing_strategy == "zero":
            df[numeric_cols] = df[numeric_cols].fillna(0)
            audit.append("Filled missing numeric values with 0")

        # ── Outlier removal (IQR method) ───────────────────────────────────
        if req.remove_outliers and numeric_cols:
            mask = pd.Series([True] * len(df), index=df.index)
            for col in numeric_cols:
                q1, q3 = df[col].quantile(0.25), df[col].quantile(0.75)
                iqr = q3 - q1
                col_mask = (df[col] >= q1 - 1.5 * iqr) & (df[col] <= q3 + 1.5 * iqr)
                removed = int((~col_mask).sum())
                if removed:
                    audit.append(f"Removed {removed} outlier rows in '{col}' (IQR method)")
                mask &= col_mask
            df = df[mask]

        # ── Z-score normalisation ──────────────────────────────────────────
        if req.normalize:
            for col in numeric_cols:
                std = float(df[col].std())
                if std > 0:
                    df[col] = (df[col] - df[col].mean()) / std
                    audit.append(f"Z-score normalised '{col}'")

        return {
            "rows": df.to_dict("records"),
            "columns": df.columns.tolist(),
            "original_count": original_count,
            "final_count": len(df),
            "rows_removed": original_count - len(df),
            "audit": audit,
        }

    except ImportError:
        # Bare fallback: drop rows with any None/empty string
        cleaned = [r for r in req.rows if all(v not in (None, "") for v in r.values())]
        return {
            "rows": cleaned,
            "columns": list(cleaned[0].keys()) if cleaned else [],
            "original_count": len(req.rows),
            "final_count": len(cleaned),
            "rows_removed": len(req.rows) - len(cleaned),
            "audit": ["pandas unavailable — dropped rows with missing values only"],
        }
    except Exception as exc:
        raise HTTPException(status_code=422, detail=str(exc))


# ── Data transform ────────────────────────────────────────────────────────────

@router.post(
    "/data/transform",
    summary="Apply a mathematical transformation to a dataset column",
    status_code=status.HTTP_200_OK,
)
async def transform_dataset(req: DataTransformRequest) -> dict:
    if not req.rows:
        raise HTTPException(status_code=422, detail="rows must not be empty")

    try:
        import pandas as pd  # type: ignore[import]

        df = pd.DataFrame(req.rows)
        col = req.column
        if col not in df.columns:
            raise HTTPException(
                status_code=400,
                detail=f"Column '{col}' not found. Available: {df.columns.tolist()}",
            )

        new_col = f"{col}_{req.transform_type}"
        audit_msg = ""

        if req.transform_type == "log":
            df[col] = df[col].clip(lower=1e-10)
            df[new_col] = np.log(df[col].astype(float))
            audit_msg = f"ln({col}) → {new_col}"

        elif req.transform_type == "sqrt":
            df[col] = df[col].clip(lower=0)
            df[new_col] = np.sqrt(df[col].astype(float))
            audit_msg = f"sqrt({col}) → {new_col}"

        elif req.transform_type == "zscore":
            mean_v = float(df[col].mean())
            std_v = float(df[col].std())
            df[new_col] = (df[col] - mean_v) / std_v if std_v > 0 else 0.0
            audit_msg = f"zscore({col}) [μ={mean_v:.3f}, σ={std_v:.3f}] → {new_col}"

        elif req.transform_type == "minmax":
            min_v = float(df[col].min())
            max_v = float(df[col].max())
            df[new_col] = (df[col] - min_v) / (max_v - min_v) if max_v > min_v else 0.0
            audit_msg = f"minmax({col}) [{min_v:.3f}–{max_v:.3f}] → {new_col} [0,1]"

        elif req.transform_type == "group_by":
            counts = df.groupby(col).size().reset_index(name=f"{col}_count")
            df = df.merge(counts, on=col, how="left")
            new_col = f"{col}_count"
            audit_msg = f"Group-by count on '{col}' → {new_col}"

        else:
            raise HTTPException(
                status_code=400, detail=f"Unknown transform_type: '{req.transform_type}'"
            )

        return {
            "rows": df.to_dict("records"),
            "columns": df.columns.tolist(),
            "new_column": new_col,
            "transform": req.transform_type,
            "audit": audit_msg,
        }

    except HTTPException:
        raise
    except Exception as exc:
        raise HTTPException(status_code=422, detail=str(exc))


# ── Helper ────────────────────────────────────────────────────────────────────

def _coerce(val: str) -> str | int | float | None:
    if not val or val.strip() == "":
        return None
    s = val.strip()
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        pass
    return s
