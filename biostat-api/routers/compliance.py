"""
Regulatory compliance endpoints for NuPhorm Biostatistics Platform.

  POST /api/v1/compliance/audit          — log a user action
  GET  /api/v1/compliance/audit          — retrieve audit trail
  POST /api/v1/compliance/validate       — run GLP/GCP validation checks
  POST /api/v1/compliance/report/excel   — generate ICH E3-style TLF Excel workbook
  POST /api/v1/compliance/export-pdf     — embed a chart image in a regulatory PDF
  POST /api/v1/compliance/ectd-package   — assemble minimal eCTD-compatible ZIP

All data persist to a local SQLite DB (compliance.db) so the audit trail
survives server restarts.
"""
from __future__ import annotations

import io
import json
import logging
import os
import sqlite3
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from fastapi import APIRouter, HTTPException, status
from fastapi.responses import Response, StreamingResponse
from pydantic import BaseModel

logger = logging.getLogger(__name__)
router = APIRouter()

# ── SQLite setup ──────────────────────────────────────────────────────────────

DB_PATH = Path(__file__).parent.parent / "compliance.db"


def _get_db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("""
        CREATE TABLE IF NOT EXISTS audit_log (
            id        INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT    NOT NULL,
            user      TEXT    NOT NULL DEFAULT 'analyst',
            action    TEXT    NOT NULL,
            category  TEXT    NOT NULL DEFAULT 'GENERAL',
            details   TEXT,
            study_id  TEXT
        )
    """)
    conn.commit()
    return conn


# ── Input models ──────────────────────────────────────────────────────────────

class AuditEntry(BaseModel):
    action: str
    category: str = "GENERAL"   # UPLOAD | ANALYSIS | EXPORT | SETTINGS | COMPLIANCE
    details: str = ""
    user: str = "analyst"
    study_id: str = ""


class ValidationRequest(BaseModel):
    data_rows: list[dict[str, Any]] = []
    study_id: str = ""
    checks: list[str] = []   # empty → run all


class TLFReportRequest(BaseModel):
    title: str = "NuPhorm Biostatistics Report"
    study_id: str = ""
    protocol_number: str = ""
    sponsor: str = "NuPhorm"
    indication: str = ""
    tables: list[dict[str, Any]] = []
    listings: list[dict[str, Any]] = []
    figures: list[dict[str, Any]] = []
    include_audit: bool = True


class PDFExportRequest(BaseModel):
    image_base64: str
    chart_type: str = "analysis"
    title: str = "NuPhorm Analysis Figure"
    study_id: str = ""
    protocol_number: str = ""


# ── Audit trail ───────────────────────────────────────────────────────────────

@router.post("/compliance/audit", summary="Log a user action to the audit trail")
async def log_audit(entry: AuditEntry) -> dict:
    ts = datetime.now(timezone.utc).isoformat()
    with _get_db() as conn:
        conn.execute(
            "INSERT INTO audit_log (timestamp, user, action, category, details, study_id) VALUES (?,?,?,?,?,?)",
            (ts, entry.user, entry.action, entry.category, entry.details, entry.study_id),
        )
    return {"logged": True, "timestamp": ts}


@router.get("/compliance/audit", summary="Retrieve audit trail")
async def get_audit(
    study_id: str = "",
    limit: int = 200,
    offset: int = 0,
) -> dict:
    with _get_db() as conn:
        if study_id:
            rows = conn.execute(
                "SELECT * FROM audit_log WHERE study_id=? ORDER BY id DESC LIMIT ? OFFSET ?",
                (study_id, limit, offset),
            ).fetchall()
            total = conn.execute("SELECT COUNT(*) FROM audit_log WHERE study_id=?", (study_id,)).fetchone()[0]
        else:
            rows = conn.execute(
                "SELECT * FROM audit_log ORDER BY id DESC LIMIT ? OFFSET ?",
                (limit, offset),
            ).fetchall()
            total = conn.execute("SELECT COUNT(*) FROM audit_log").fetchone()[0]
    return {
        "entries": [dict(r) for r in rows],
        "total": total,
        "limit": limit,
        "offset": offset,
    }


# ── GLP/GCP validation ────────────────────────────────────────────────────────

GLP_GCP_CHECKS = {
    "DATA_COMPLETENESS": "No records with all-null/empty values",
    "SUBJECT_ID_UNIQUE": "USUBJID / subject identifier is unique per record",
    "DATE_CONSISTENCY": "Date fields are in ISO-8601 or common date format",
    "NUMERIC_RANGE": "Numeric values are within plausible biological range (no extreme outliers > 10 SD)",
    "REQUIRED_COLUMNS": "Dataset contains at minimum: USUBJID, STUDYID or equivalent",
    "TRACEABILITY": "At least one data source has been registered in the audit trail",
    "DUPLICATE_RECORDS": "No exact duplicate records",
    "MISSING_RATE": "Missing rate per column does not exceed 20%",
}


@router.post("/compliance/validate", summary="Run GLP/GCP validation checks on a dataset")
async def validate_dataset(req: ValidationRequest) -> dict:
    try:
        import pandas as pd  # type: ignore[import]
        import numpy as np

        df = pd.DataFrame(req.data_rows) if req.data_rows else pd.DataFrame()
        checks_to_run = set(req.checks) if req.checks else set(GLP_GCP_CHECKS.keys())
        results: list[dict] = []

        def _check(key: str, passed: bool, detail: str, severity: str = "ERROR"):
            if key in checks_to_run:
                results.append({"check": key, "description": GLP_GCP_CHECKS.get(key, key), "passed": passed, "detail": detail, "severity": severity if not passed else "OK"})

        if df.empty:
            return {"passed": False, "checks": [{"check": "DATA_COMPLETENESS", "description": "Dataset is empty", "passed": False, "detail": "No rows provided", "severity": "ERROR"}]}

        # DATA_COMPLETENESS
        all_null = df.isnull().all(axis=1).sum()
        _check("DATA_COMPLETENESS", all_null == 0, f"{all_null} records have all-null values")

        # SUBJECT_ID_UNIQUE
        id_col = next((c for c in df.columns if c.upper() in {"USUBJID", "SUBJID", "PATIENTID", "SUBJECT_ID", "ID"}), None)
        if id_col:
            dupes = df[id_col].duplicated().sum()
            _check("SUBJECT_ID_UNIQUE", dupes == 0, f"{dupes} duplicate subject IDs in '{id_col}'")
        else:
            _check("SUBJECT_ID_UNIQUE", False, "No subject identifier column found (expected USUBJID, ID, etc.)", "WARNING")

        # REQUIRED_COLUMNS
        has_req = any(c.upper() in {"USUBJID", "SUBJID", "STUDYID", "ID", "SUBJECT_ID"} for c in df.columns)
        _check("REQUIRED_COLUMNS", has_req, "Missing USUBJID/STUDYID column" if not has_req else "Required columns present")

        # DUPLICATE_RECORDS
        dup_rows = df.duplicated().sum()
        _check("DUPLICATE_RECORDS", dup_rows == 0, f"{dup_rows} exact duplicate rows", "WARNING")

        # MISSING_RATE
        high_miss = [(c, round(df[c].isnull().mean() * 100, 1)) for c in df.columns if df[c].isnull().mean() > 0.20]
        _check("MISSING_RATE", len(high_miss) == 0, ("Columns >20% missing: " + ", ".join(f"{c}({p}%)" for c, p in high_miss)) if high_miss else "All columns within threshold", "WARNING")

        # NUMERIC_RANGE
        num_cols = df.select_dtypes(include=[np.number]).columns
        extreme = []
        for col in num_cols:
            mu, sd = df[col].mean(), df[col].std()
            if sd > 0 and (np.abs(df[col] - mu) > 10 * sd).any():
                extreme.append(col)
        _check("NUMERIC_RANGE", len(extreme) == 0, f"Extreme outliers (>10 SD) in: {', '.join(extreme)}" if extreme else "No extreme outliers detected", "WARNING")

        # DATE_CONSISTENCY — check string columns that look like dates
        date_cols = [c for c in df.select_dtypes(include=["object"]).columns if "date" in c.lower() or "dt" in c.lower()]
        bad_dates = []
        for col in date_cols:
            try:
                pd.to_datetime(df[col].dropna(), errors="raise")
            except Exception:
                bad_dates.append(col)
        _check("DATE_CONSISTENCY", len(bad_dates) == 0, f"Unparseable date columns: {', '.join(bad_dates)}" if bad_dates else "All date columns parse correctly", "WARNING")

        # TRACEABILITY — check audit log for uploads
        with _get_db() as conn:
            audit_count = conn.execute(
                "SELECT COUNT(*) FROM audit_log WHERE category='UPLOAD'" + (" AND study_id=?" if req.study_id else ""),
                (req.study_id,) if req.study_id else (),
            ).fetchone()[0]
        _check("TRACEABILITY", audit_count > 0, f"{audit_count} upload events in audit trail" if audit_count else "No UPLOAD events found in audit trail", "WARNING")

        passed = all(r["severity"] in ("OK", "WARNING") for r in results)
        score = round(100 * sum(1 for r in results if r["severity"] == "OK") / max(len(results), 1))
        return {"passed": passed, "compliance_score": score, "checks": results, "summary": f"{sum(1 for r in results if r['severity'] == 'OK')}/{len(results)} checks passed"}

    except ImportError:
        return {"passed": False, "checks": [{"check": "ENVIRONMENT", "description": "pandas not installed", "passed": False, "detail": "Run: pip install pandas", "severity": "ERROR"}]}
    except Exception as exc:
        raise HTTPException(status_code=422, detail=str(exc))


# ── TLF Excel report ──────────────────────────────────────────────────────────

@router.post(
    "/compliance/report/excel",
    summary="Generate ICH E3-style TLF Excel workbook",
    response_class=Response,
)
async def generate_tlf_excel(req: TLFReportRequest) -> Response:
    try:
        import openpyxl  # type: ignore[import]
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(status_code=501, detail="openpyxl not installed. Run: pip install openpyxl")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Colour palette ───────────────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="1E3A5F")
    SUB_FILL    = PatternFill("solid", fgColor="D6E4F0")
    TITLE_FONT  = Font(name="Calibri", size=12, bold=True, color="FFFFFF")
    HDR_FONT    = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    BODY_FONT   = Font(name="Calibri", size=10)
    BOLD        = Font(name="Calibri", size=10, bold=True)
    THIN        = Side(style="thin", color="B0B0B0")
    BORDER      = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CENTER      = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT        = Alignment(horizontal="left", vertical="center", wrap_text=True)

    def _header_row(ws, row: int, cols: list[str]):
        for c, val in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.fill = HEADER_FILL
            cell.font = HDR_FONT
            cell.alignment = CENTER
            cell.border = BORDER

    def _cover_sheet():
        ws = wb.create_sheet("Cover")
        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 60
        fields = [
            ("Report Title", req.title),
            ("Study ID / Protocol", req.study_id or req.protocol_number or "—"),
            ("Sponsor", req.sponsor),
            ("Indication", req.indication or "—"),
            ("Generated By", "NuPhorm Biostatistics Platform"),
            ("Generated On", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")),
            ("Standard", "ICH E3 / ICH M8 (eCTD)"),
        ]
        ws.merge_cells("A1:B1")
        title_cell = ws["A1"]
        title_cell.value = "CLINICAL STUDY REPORT — TLF COMPILATION"
        title_cell.fill = HEADER_FILL
        title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
        title_cell.alignment = CENTER
        ws.row_dimensions[1].height = 30

        for i, (label, val) in enumerate(fields, 2):
            lc = ws.cell(row=i, column=1, value=label)
            lc.font = BOLD
            lc.fill = SUB_FILL
            lc.alignment = LEFT
            lc.border = BORDER
            vc = ws.cell(row=i, column=2, value=val)
            vc.font = BODY_FONT
            vc.alignment = LEFT
            vc.border = BORDER

    def _table_sheet(idx: int, tbl: dict):
        name = tbl.get("title", f"Table {idx + 1}")[:31]
        ws = wb.create_sheet(f"T{idx+1}_{name[:25]}")
        headers: list[str] = tbl.get("headers", [])
        rows: list[list] = tbl.get("rows", [])
        footnotes: list[str] = tbl.get("footnotes", [])

        # Title block
        ws.merge_cells(f"A1:{get_column_letter(max(len(headers), 1))}1")
        t = ws["A1"]
        t.value = name
        t.font = Font(name="Calibri", size=11, bold=True)
        t.alignment = CENTER
        ws.row_dimensions[1].height = 22

        if headers:
            _header_row(ws, 2, headers)
            for ri, row in enumerate(rows, 3):
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font = BODY_FONT
                    cell.border = BORDER
                    cell.alignment = LEFT
                    if ri % 2 == 0:
                        cell.fill = PatternFill("solid", fgColor="F8FAFC")

        # Auto column width
        for ci, hdr in enumerate(headers, 1):
            col_vals = [str(hdr)] + [str(r[ci - 1]) if ci - 1 < len(r) else "" for r in rows]
            ws.column_dimensions[get_column_letter(ci)].width = min(40, max(12, max(len(v) for v in col_vals) + 2))

        # Footnotes
        if footnotes:
            fn_row = (len(rows) + 4) if rows else 4
            for i, fn in enumerate(footnotes):
                ws.cell(row=fn_row + i, column=1, value=f"[{i+1}] {fn}").font = Font(name="Calibri", size=9, italic=True)

    def _listing_sheet(idx: int, lst: dict):
        # Listings share the same structure as tables; just different sheet prefix
        name = lst.get("title", f"Listing {idx + 1}")[:31]
        ws = wb.create_sheet(f"L{idx+1}_{name[:25]}")
        headers = lst.get("headers", [])
        rows    = lst.get("rows", [])

        ws.merge_cells(f"A1:{get_column_letter(max(len(headers), 1))}1")
        t = ws["A1"]
        t.value = name
        t.font = Font(name="Calibri", size=11, bold=True)
        t.alignment = CENTER
        ws.row_dimensions[1].height = 22

        if headers:
            _header_row(ws, 2, headers)
            for ri, row in enumerate(rows, 3):
                for ci, val in enumerate(row, 1):
                    cell = ws.cell(row=ri, column=ci, value=val)
                    cell.font = Font(name="Calibri", size=9)
                    cell.border = BORDER
                    cell.alignment = LEFT

    def _audit_sheet():
        ws = wb.create_sheet("Audit Trail")
        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 50
        _header_row(ws, 1, ["Timestamp (UTC)", "User", "Action", "Category", "Details"])
        with _get_db() as conn:
            rows = conn.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT 500").fetchall()
        for ri, row in enumerate(rows, 2):
            for ci, val in enumerate([row["timestamp"], row["user"], row["action"], row["category"], row["details"]], 1):
                cell = ws.cell(row=ri, column=ci, value=val)
                cell.font = Font(name="Calibri", size=9)
                cell.border = BORDER
                if ri % 2 == 0:
                    cell.fill = PatternFill("solid", fgColor="F8FAFC")

    # Build workbook
    _cover_sheet()
    for i, tbl in enumerate(req.tables):
        _table_sheet(i, tbl)
    for i, lst in enumerate(req.listings):
        _listing_sheet(i, lst)
    if req.include_audit:
        _audit_sheet()

    # Log the export
    with _get_db() as conn:
        conn.execute(
            "INSERT INTO audit_log (timestamp, user, action, category, details, study_id) VALUES (?,?,?,?,?,?)",
            (datetime.now(timezone.utc).isoformat(), "analyst", "REPORT_GENERATED", "EXPORT", f"TLF Excel: {req.title}", req.study_id),
        )

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"TLF_{req.study_id or 'report'}_{datetime.now().strftime('%Y%m%d')}.xlsx"
    return Response(
        content=buf.read(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── PDF figure export ─────────────────────────────────────────────────────────

@router.post(
    "/compliance/export-pdf",
    summary="Embed a chart image in a regulatory-formatted PDF",
    response_class=Response,
)
async def export_chart_pdf(req: PDFExportRequest) -> Response:
    import base64
    try:
        from reportlab.lib.pagesizes import A4, landscape  # type: ignore[import]
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Image, Paragraph, Spacer, Table, TableStyle
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib import colors
    except ImportError:
        # Fallback: return the raw PNG
        img_bytes = base64.b64decode(req.image_base64)
        return Response(content=img_bytes, media_type="image/png", headers={"Content-Disposition": f'attachment; filename="figure_{req.chart_type}.png"'})

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=1.5 * cm, bottomMargin=1.5 * cm, leftMargin=2 * cm, rightMargin=2 * cm)
    styles = getSampleStyleSheet()
    story = []

    # Header
    header_style = ParagraphStyle("hdr", parent=styles["Normal"], fontSize=9, textColor=colors.HexColor("#1E3A5F"), spaceAfter=4)
    story.append(Paragraph(f"<b>NuPhorm Biostatistics Platform</b>  ·  {req.sponsor if hasattr(req, 'sponsor') else 'NuPhorm'}", header_style))
    if req.study_id:
        story.append(Paragraph(f"Study ID: {req.study_id}" + (f"  ·  Protocol: {req.protocol_number}" if req.protocol_number else ""), header_style))

    # Title
    title_style = ParagraphStyle("title", parent=styles["Heading2"], fontSize=14, spaceAfter=8, textColor=colors.HexColor("#1E3A5F"))
    story.append(Paragraph(req.title, title_style))

    # Chart image
    img_bytes = base64.b64decode(req.image_base64)
    img_buf = io.BytesIO(img_bytes)
    img = Image(img_buf, width=22 * cm, height=12 * cm, hAlign="CENTER")
    story.append(img)
    story.append(Spacer(1, 0.5 * cm))

    # Footer
    footer_style = ParagraphStyle("footer", parent=styles["Normal"], fontSize=8, textColor=colors.HexColor("#6B7280"))
    story.append(Paragraph(f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M UTC')}  ·  Pharma-Ready Output  ·  NuPhorm Platform", footer_style))

    doc.build(story)
    buf.seek(0)
    filename = f"NuPhorm_{req.chart_type}_{datetime.now().strftime('%Y%m%d')}.pdf"
    return Response(
        content=buf.read(),
        media_type="application/pdf",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


# ── eCTD package ──────────────────────────────────────────────────────────────

@router.post(
    "/compliance/ectd-package",
    summary="Assemble a minimal eCTD-compatible ZIP archive",
    response_class=Response,
)
async def create_ectd_package(req: TLFReportRequest) -> Response:
    """Creates a ZIP with the ICH eCTD folder structure (m5/53/clinical-studies)."""
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        base = f"eCTD_{req.study_id or 'study'}/m5/53/5351/clinical-studies/"

        # Placeholder index.xml
        index_xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<ich:ectd xmlns:ich="http://www.ich.org/ectd" version="3.2.2">
  <header>
    <study-id>{req.study_id}</study-id>
    <title>{req.title}</title>
    <sponsor>{req.sponsor}</sponsor>
    <date>{datetime.now(timezone.utc).date()}</date>
  </header>
  <m5>
    <s53><s5351><clinical-studies>
      <!-- TLF reports generated by NuPhorm Biostatistics Platform -->
    </clinical-studies></s5351></s53>
  </m5>
</ich:ectd>"""
        zf.writestr(f"{base}index.xml", index_xml)

        # Audit trail JSON
        with _get_db() as conn:
            audit = [dict(r) for r in conn.execute("SELECT * FROM audit_log ORDER BY id DESC LIMIT 1000").fetchall()]
        zf.writestr(f"{base}audit_trail.json", json.dumps(audit, indent=2, default=str))

        # Manifest
        manifest = {
            "ectd_version": "3.2.2",
            "generated_by": "NuPhorm Biostatistics Platform",
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "study_id": req.study_id,
            "title": req.title,
            "sponsor": req.sponsor,
            "standard": "ICH M8 / eCTD v3.2",
            "contents": ["index.xml", "audit_trail.json"],
        }
        zf.writestr(f"{base}manifest.json", json.dumps(manifest, indent=2))

    buf.seek(0)
    filename = f"eCTD_{req.study_id or 'package'}_{datetime.now().strftime('%Y%m%d')}.zip"
    return Response(
        content=buf.read(),
        media_type="application/zip",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
